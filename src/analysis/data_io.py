import os, sys, glob
import pandas as pd
import numpy as np
from datetime import datetime

src_dir = os.path.abspath(os.path.join(os.pardir,'src'))
sys.path[0] = src_dir

def parse_data_IPIP(path_filename,path_output,Q=120):
    '''Read the original IPIP-dataset into csv-files that are easily readable as pandas-dataframes
    IN:
    - path_filename, str, path+filename of IPIP-dataset
    - path_output, str, where to save the csv-file
    - Q, int, number of questions in respective dataset (default=120)
    OUT:
    - nothing, files are saved in path_output
    '''
    list_questions = [ 'Q_%s'%(h) for h in np.arange(Q)+1]
    list_metadata = ['case','gender','age','country','year','datetime']
    list_columns = list_metadata+list_questions

    with open(path_filename,'r') as f:
        x=f.readlines()

    list_case = []
    list_gender = []
    list_age = []
    list_country = []
    list_year = []
    list_datetime = []
    list_answers = []

    if Q == 300:
        i_off = 2
    else:
        i_off = 0

    for item in x:
        case = int(item[i_off+0:i_off+6].rstrip())
        gender = int(item[i_off+6])
        age = int(item[i_off+7:i_off+9])
        ss = int(item[i_off+9:i_off+11])
        mm = int(item[i_off+11:i_off+13])
        hh = int(item[i_off+13:i_off+15])
        day =  int(item[i_off+15:i_off+17])
        month =  int(item[i_off+17:i_off+19])
        year =  int(item[i_off+19:i_off+22])+1900
        t_datetime = str(datetime(year,month+1,day,hh,mm,ss))
        # ## note that 0 means missing answer
        if Q == 120:
            country = item[i_off+22:i_off+31].rstrip()
            answers = [int(h) for h in item[i_off+31:-1]]
        elif Q == 300:
            country = item[i_off+22:i_off+33].rstrip()
            answers = [int(h) for h in item[i_off+33:-1]]
        else:
            print('invalid number for number of questions!!!')
            break 
        
        list_case += [case]
        list_gender += [gender]
        list_age += [age]
        list_country += [country]
        list_year += [year]
        list_datetime += [t_datetime]
        list_answers += [answers]

    ## make dataframe
    df = pd.DataFrame(columns=list_columns)
    df['case'] = list_case
    df['gender'] = list_gender
    df['age'] = list_age
    df['country'] = list_country
    df['year'] = list_year
    df['datetime'] = list_datetime

    for i_q,q in enumerate(list_questions):
        df[q] = [  h[i_q] for h in list_answers  ]
    ## replace non-answers (0) with np.nan
    df=df.replace(0,np.nan)

    fname_save = 'df_filter_'+''.join(path_filename.split('/')[-1].split('.')[:-1])+'.csv'
    df.to_csv(os.path.join(path_output,fname_save))
    return

def get_data_array(path_filename_df_filter_csv,Q=120,shift=True,nan_remove=False,nan_replace = 0.0, return_case=False):
    '''
    Retrieve the csv-file dataframes as an array
    IN:
    - path_filename_df_filter_csv, str: filename to csv-dataframe
    - Q, int, number of questions (default=120)
    - shift, bool (default=True), whether to transform from 1,...,5 to -2,...,2
    - nan_remove, bool (default=False). whether to remove entries that do not contain all answers. 
        if false: repalce answer by nan_replace
    - nan_replace, float (default:0.0): value that replaces missing entries
    - return_case, bool (default=False): whether to return indices (id-numbers) as additional list
    OUT:
    - arr, shape: samples (persons) x questions (features)
    '''
    df = pd.read_csv(path_filename_df_filter_csv,index_col=0)
    list_questions = [ 'Q_%s'%(h) for h in np.arange(Q)+1]

    ## if shift == True:
    ## answer-values={1,2,3,4,5} |--> {-2,-1,0,1,2} 
    if shift == True:
        df[list_questions] = df[list_questions]-3

    ## if nan_remove==True: remove answers with nan
    ## if False , we give value nan_replace
    if nan_remove == True:
        df=df.dropna()
    else:
        df=df.replace(np.nan,nan_replace)
    ## if return_case == True, we also return the list of cases we keep
    if return_case == True:
        return np.array(df[list_questions]).astype(int),list(df['case'].values)
    else:
        return np.array(df[list_questions]).astype(int)
    

def read_domains_IPIP120(FILENAME,return_facet_names=False):
    from openpyxl import load_workbook

    wb = load_workbook(FILENAME, read_only=True)
    ws = wb['IPIP-NEO-ItemKey'] # ws is now an IterableWorksheet

    domains = []
    domains_sgn = []
    facet_names = []
    for i_row,row in enumerate(ws.rows):
        if i_row == 0:
            pass
        else:
            cont_ = []
            for i_cell,cell in enumerate(row):
                cont_ += [cell.value]
            if cont_[1] == None:
                break
            domains += [cont_[3]]

            sgn_ = 0
            if cont_[2][0]=='+':
                sgn_ = 1
            if cont_[2][0]=='-':
                sgn_ = -1
            domains_sgn += [sgn_]
            facet_names += [cont_[4]]

    if return_facet_names == False:
        return domains,np.array(domains_sgn)
    else:
        return domains,np.array(domains_sgn),facet_names
def read_domains_IPIP300(FILENAME,return_facet_names=False):
    from openpyxl import load_workbook

    wb = load_workbook(FILENAME, read_only=True)
    ws = wb['IPIP-NEO-ItemKey'] # ws is now an IterableWorksheet

    domains = []
    domains_sgn = []
    index_q = []
    facet_names = []
    for i_row,row in enumerate(ws.rows):
        if i_row == 0:
            pass
        else:
            cont_ = []
            for i_cell,cell in enumerate(row):
                cont_ += [cell.value]
            if cont_[0] == None:
                break
            domains += [cont_[3]]

            sgn_ = 0
            if cont_[2][0]=='+':
                sgn_ = 1
            if cont_[2][0]=='-':
                sgn_ = -1
            domains_sgn += [sgn_]
            facet_names += [cont_[4]]

            index_q += [int(cont_[0])]
    index_q_sort = np.argsort(np.array(index_q))
    domains_sort = [domains[i] for i in index_q_sort]
    domains_sgn_sort = [domains_sgn[i] for i in index_q_sort]
    facet_names_sort = [facet_names[i] for i in index_q_sort]
    if return_facet_names == False:
        return domains_sort,np.array(domains_sgn_sort)
    else:
        return domains_sort,np.array(domains_sgn_sort),np.array(facet_names_sort)

def score_df_pq_to_pd(df_pq,df_qd,reverse = True, R=None, zscore = False,dropna = True):
    list_doms = df_qd.columns
    df_pd = pd.DataFrame(index=df_pq.index,columns = list_doms)
    ## the maximum answer - needed for reversing the answers
    if R == None:
        R = np.nanmax(np.array(df_pq))+1
    for d in list_doms:
        ## get questions and scoring for each domain
        s=df_qd[d].dropna()
        list_q_tmp = s.keys()
        list_s_tmp = s[list_q_tmp]

        ## get the subset of questions that everyone answered
        ## reverse the answer if necessar
        H = df_pq.loc[:,list_q_tmp]
        for i_q,q in enumerate(list_q_tmp):
            h = s[q]
            if h<0 and reverse == True:
                H[q] = R - H[q]
        df_pd[d] = H.mean(axis=1)
    ## Drop people for which we do not have a score on everything
    if dropna==True:
        df_pd = df_pd.dropna(how='any',axis=0)
    
    if zscore == True:
        df_pd = (df_pd-df_pd.mean())/df_pd.std()
    return df_pd